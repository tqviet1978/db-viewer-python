"""Excel export using openpyxl — ported from PHP ExcelHelper."""

from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def export_to_excel(
    rows: list[dict],
    columns: list[str],
    column_titles: list[str],
    decimal_columns: list[str],
    text_columns: list[str],
    summable_columns: list[str],
    align_center_columns: list[str],
    sheet_separation_column: str,
    column_widths: list[str | int],
) -> BytesIO:
    """
    Generate an Excel file and return as BytesIO.

    - First column is always '#' (row index)
    - Apply column_titles as header row
    - Format decimal_columns with 6 decimal places
    - Format text_columns as text
    - Add SUM formulas for summable_columns at bottom
    - Center-align align_center_columns
    - Set column widths
    - If sheet_separation_column: group rows by that column value into separate sheets + All
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Parse column widths (prepend index column width)
    widths = _parse_widths(column_widths, len(columns))

    if not sheet_separation_column:
        ws = wb.create_sheet("DATA")
        _write_sheet(ws, rows, columns, column_titles, decimal_columns,
                     text_columns, summable_columns, align_center_columns, widths)
    else:
        # Group rows
        grouped: dict[str, list[dict]] = {}
        for row in rows:
            key = str(row.get(sheet_separation_column, "") or "Unknown")
            grouped.setdefault(key, []).append(row)

        for sheet_key, sheet_rows in grouped.items():
            ws = wb.create_sheet(sheet_key[:31])  # Excel max sheet name length
            _write_sheet(ws, sheet_rows, columns, column_titles, decimal_columns,
                         text_columns, summable_columns, align_center_columns, widths)

        # All sheet
        ws_all = wb.create_sheet("All")
        _write_sheet(ws_all, rows, columns, column_titles, decimal_columns,
                     text_columns, summable_columns, align_center_columns, widths)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _parse_widths(raw: list, n_columns: int) -> list[int]:
    """Parse column widths list, prepend index column width 8."""
    widths = []
    for w in raw:
        try:
            widths.append(int(w))
        except (ValueError, TypeError):
            widths.append(15)

    # Prepend index (#) column width
    if len(widths) <= n_columns:
        widths.insert(0, 8)

    return widths


def _write_sheet(
    ws,
    rows: list[dict],
    columns: list[str],
    column_titles: list[str],
    decimal_columns: list[str],
    text_columns: list[str],
    summable_columns: list[str],
    align_center_columns: list[str],
    widths: list[int],
) -> None:
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="F6F8FA")
    center_align = Alignment(horizontal="center")

    # Build title list: first col is #
    all_cols = ["#"] + columns
    titles = ["#"] + (column_titles if column_titles else columns)

    # Write header
    for ci, title in enumerate(titles, start=1):
        cell = ws.cell(row=1, column=ci, value=title)
        cell.font = header_font
        cell.fill = header_fill
        if all_cols[ci - 1] in align_center_columns or ci == 1:
            cell.alignment = center_align

    # Write data rows
    for ri, row in enumerate(rows, start=2):
        ws.cell(row=ri, column=1, value=ri - 1)  # index

        for ci, col in enumerate(columns, start=2):
            value = row.get(col)

            if col in decimal_columns:
                try:
                    value = float(value)
                    cell = ws.cell(row=ri, column=ci, value=value)
                    cell.number_format = "#,##0.000000"
                except (TypeError, ValueError):
                    ws.cell(row=ri, column=ci, value=value)
            elif col in text_columns:
                cell = ws.cell(row=ri, column=ci, value=str(value) if value is not None else "")
                cell.number_format = "@"
            else:
                ws.cell(row=ri, column=ci, value=value)

            if col in align_center_columns:
                ws.cell(row=ri, column=ci).alignment = center_align

    # SUM row
    if rows and summable_columns:
        sum_row = len(rows) + 2
        ws.cell(row=sum_row, column=1, value="Total").font = header_font
        for ci, col in enumerate(columns, start=2):
            if col in summable_columns:
                col_letter = get_column_letter(ci)
                last_data_row = len(rows) + 1
                ws.cell(row=sum_row, column=ci, value=f"=SUM({col_letter}2:{col_letter}{last_data_row})")
                ws.cell(row=sum_row, column=ci).font = header_font

    # Column widths
    for ci, col_name in enumerate(all_cols, start=1):
        col_letter = get_column_letter(ci)
        if ci - 1 < len(widths):
            ws.column_dimensions[col_letter].width = widths[ci - 1]
        else:
            ws.column_dimensions[col_letter].width = 15
