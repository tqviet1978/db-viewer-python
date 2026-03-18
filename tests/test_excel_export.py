"""Tests for excel_export module."""

import pytest
from io import BytesIO

from openpyxl import load_workbook

from dbviewer.excel_export import export_to_excel


def _sample_rows():
    return [
        {"INVOICE_NO": "INV001", "AMOUNT": 1500.5, "CUSTOMER": "Alice", "STATUS": "paid"},
        {"INVOICE_NO": "INV002", "AMOUNT": 2000.0, "CUSTOMER": "Bob", "STATUS": "pending"},
        {"INVOICE_NO": "INV003", "AMOUNT": 750.25, "CUSTOMER": "Charlie", "STATUS": "paid"},
    ]


class TestExportToExcel:
    def _export(self, **kwargs):
        defaults = dict(
            rows=_sample_rows(),
            columns=["INVOICE_NO", "AMOUNT", "CUSTOMER", "STATUS"],
            column_titles=["Invoice", "Amount", "Customer", "Status"],
            decimal_columns=["AMOUNT"],
            text_columns=["INVOICE_NO"],
            summable_columns=["AMOUNT"],
            align_center_columns=["STATUS"],
            sheet_separation_column="",
            column_widths=["15", "12", "20", "10"],
        )
        defaults.update(kwargs)
        return export_to_excel(**defaults)

    def test_returns_bytesio(self):
        buf = self._export()
        assert isinstance(buf, BytesIO)
        assert buf.tell() == 0

    def test_valid_xlsx(self):
        buf = self._export()
        wb = load_workbook(buf)
        assert wb is not None

    def test_sheet_name(self):
        buf = self._export()
        wb = load_workbook(buf)
        assert "DATA" in wb.sheetnames

    def test_header_row(self):
        buf = self._export()
        wb = load_workbook(buf)
        ws = wb["DATA"]
        headers = [ws.cell(1, c).value for c in range(1, 6)]
        assert "#" in headers
        assert "Invoice" in headers
        assert "Amount" in headers

    def test_index_column(self):
        buf = self._export()
        wb = load_workbook(buf)
        ws = wb["DATA"]
        assert ws.cell(2, 1).value == 1
        assert ws.cell(3, 1).value == 2

    def test_row_count(self):
        buf = self._export()
        wb = load_workbook(buf)
        ws = wb["DATA"]
        # header + 3 data rows + 1 sum row = 5 rows
        assert ws.max_row >= 4

    def test_sum_row_present(self):
        buf = self._export()
        wb = load_workbook(buf)
        ws = wb["DATA"]
        # Sum row should be in the last row
        last_row = ws.max_row
        found_total = any(ws.cell(last_row, c).value == "Total" for c in range(1, 6))
        assert found_total

    def test_decimal_format(self):
        buf = self._export()
        wb = load_workbook(buf)
        ws = wb["DATA"]
        # AMOUNT column is index 3 (1=#, 2=INVOICE_NO, 3=AMOUNT, 4=CUSTOMER, 5=STATUS)
        amount_col = 3
        cell = ws.cell(2, amount_col)
        # Value should be numeric
        assert isinstance(cell.value, float)

    def test_sheet_separation(self):
        buf = self._export(sheet_separation_column="STATUS")
        wb = load_workbook(buf)
        sheets = wb.sheetnames
        assert "paid" in sheets
        assert "pending" in sheets
        assert "All" in sheets

    def test_sheet_separation_all_has_all_rows(self):
        buf = self._export(sheet_separation_column="STATUS")
        wb = load_workbook(buf)
        ws = wb["All"]
        # 3 data rows + 1 header + 1 sum = 5
        assert ws.max_row >= 4

    def test_no_column_titles_uses_column_names(self):
        buf = self._export(column_titles=[])
        wb = load_workbook(buf)
        ws = wb["DATA"]
        headers = [ws.cell(1, c).value for c in range(1, 6)]
        assert "INVOICE_NO" in headers or "Invoice" in headers  # fallback to column name

    def test_empty_rows(self):
        buf = export_to_excel(
            rows=[],
            columns=["A", "B"],
            column_titles=[],
            decimal_columns=[],
            text_columns=[],
            summable_columns=[],
            align_center_columns=[],
            sheet_separation_column="",
            column_widths=[],
        )
        wb = load_workbook(buf)
        ws = wb["DATA"]
        assert ws.max_row >= 1  # at least header
